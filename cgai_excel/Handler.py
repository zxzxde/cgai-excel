# -*- coding:utf-8 -*-
"""
步骤：
1. 先从 xl\worksheets路径对应的sheet1.xml来获取 DISPIMG所对应的cell格子坐标
 c r = 'B9' 以及 图片名称 ; aa2 &
 <c r="B9" t="str">
                <f>_xlfn.DISPIMG(&quot;aa2&quot;,1)</f>
                <v>=DISPIMG(&quot;aa2&quot;,1)</v>

2. 再从xl\drawings\drawing1.xml 或者xl\cellimages.xml 中进行读取，这里需要进行判定，
因为有两个位置，可能是Excel的版本问题
通过 name 来查询 所在 xdr:pic 以及该下的 xdr:blipFill里的 embed =''
        <xdr:pic>
            <xdr:nvPicPr>
                <xdr:cNvPr id="3" name="A21" descr="A21"/>
                <xdr:cNvPicPr>
                    <a:picLocks noChangeAspect="1"/>
                </xdr:cNvPicPr>
            </xdr:nvPicPr>
            <xdr:blipFill>
                <a:blip r:embed="rId2"/>
            </xdr:blipFill>

"""


from openpyxl import Workbook,load_workbook
from openpyxl.reader.drawings import  find_images
# from  zipfile import ZipFile
import os
import zipfile
import re
from openpyxl.packaging.relationship import (
    RelationshipList,
    get_dependents,
    get_rels_path,
)

import shutil
import re 
import os 
# from cgai_io import deldir,delall

ROOT = os.path.dirname(__file__)

EXTPATH = os.path.join(ROOT,'EXT')


def clearTemp():
    """
    删除解压后的Excel目录
    :return:
    """
    if os.path.exists(EXTPATH):
        shutil.rmtree(EXTPATH)

def extract_excel(excel_path,extract_dirpath=EXTPATH):
    """
    解压excel
    :param excel_path: Excel路径
    :param extract_dirpath: 解压目录
    :return:
    """
    clearTemp()
    _zip = zipfile.ZipFile(excel_path)
    _zip.extractall(extract_dirpath)


def get_image_name(excel_path,extract_dirpath=EXTPATH):
    """
    从worksheets中获取图片，仅读取sheet1
    """
    extract_excel(excel_path,extract_dirpath)
    result = []
    sheet_path = os.path.join(extract_dirpath,'xl/worksheets/sheet1.xml')
    print('sheet_path:',sheet_path)
    if os.path.exists(sheet_path):
        s = ''
        with open(sheet_path,'r',encoding='utf8') as r:
            s = r.read()
        pattern = re.compile('<c r="([A-Z]+\d+)" t="str"><f>_xlfn.DISPIMG\(&quot;(\w+)&quot;,1\)</f>')
        result = pattern.findall(s)
    return result


def get_image_path(excel_path,extract_dirpath=EXTPATH):
    """
    获取图片路径
    :param excel_path:
    :param extract_dirpath:
    :return:
    """
    global EXTPATH
    EXTPATH = extract_dirpath
    image_name_map = get_image_name(excel_path,extract_dirpath)
    # print(image_name_map)
    xml_path = ''
    xml_path1 = os.path.join(extract_dirpath,'xl/cellimages.xml')
    xml_path2 = os.path.join(extract_dirpath,'xl/drawings/drawing1.xml')
    if os.path.exists(xml_path1):
        xml_path = xml_path1
    if os.path.exists(xml_path2):
        xml_path = xml_path2
    image_id_map = {}
    media_path = os.path.join(extract_dirpath,'xl/media')
    images = os.listdir(media_path)
    if xml_path:
        s = ''
        with open(xml_path,'r',encoding='utf8') as r:
            s = r.read()
        for k,v in image_name_map:
            p = f'name="{v}.*?rId(.*?)"'
            pattern = re.compile(p, re.DOTALL)
            result = pattern.findall(s)
            rid = result[0]
            start = f'image{rid}'
            if images:
                for i in images:
                    name = os.path.splitext(i)[0]
                    if start == name:
                        image_id_map[k] = os.path.join(media_path,i)

    return image_id_map



def get_excel_data(excel_path,extract_dirpath=EXTPATH):
    """
    获取Excel数据，包含图片路径信息。图片须满足嵌入形式
    :param excel_path:
    :param extract_dirpath:
    :return:
    """
    data = {}
    wb = load_workbook(excel_path,data_only=True)
    ws = wb.active
    rows = ws.max_row
    columns = ws.max_column
    # print(ws['A1'].coordinate)
    # print(ws.cell(1,1).coordinate)
    header_row = ws[1]
    header = [h.value for h in header_row]
    # print(header)

    image_map = get_image_path(excel_path,extract_dirpath)
    data_list = []
    for r in range(1,rows+1):
        per_row_data = []
        for c in range(1,columns+1):
            cood = ws.cell(r,c).coordinate
            value = ws.cell(r,c).value
            image_path = image_map.get(cood,'')
            value = 'img:' + image_path if image_path else value
            per_row_data.append(value)
        data_list.append(per_row_data)

    data['header'] = header
    data['data_list'] = data_list
    wb.close()

    return data



# # path = r'F:\Temp\Q\AA.xlsx'
# path = r'F:\Temp\Q\excel_module.xlsx'
# extract_dirpath = r'F:\Temp\Q\Atemp'
# # extract_excel(r'C:\Temp\output\test.xlsx')
# # image_name = get_image_name()
# image_id = get_image_path(path,extract_dirpath)
# print(image_id)
# # clearTemp()

#
# s = ''
# with open(path,'r',encoding='utf8') as r:
#     s = r.read()


# pattern = re.compile('.*?<c r="([A-Z]+\d+)".*?_xlfn.DISPIMG\(&quot;(.*)?&quot.*')
# pattern = re.compile('<c r="([A-Z]+\d+)" t="str"><f>_xlfn.DISPIMG\(&quot;(\w+)&quot;,1\)</f>')

# result = pattern.findall(s)
# print(len(result))
# print(result)


# os.path.dirname()

# path = r'C:\Users\Admin\Downloads\A.xlsx'
# path = r'C:\Temp\output\test.xlsx'
# path = r'F:\Temp\Q\B.xlsx'
# zip = zipfile.ZipFile(path)
# zip.extractall(r'C:\Temp\output\test')
# zip.extractall(r'F:\Temp\Q\B')

# fz = zipfile.ZipFile(path, 'r')
# # xml_name = 'xl/drawings/drawing1.xml'
# xml_name = 'xl/drawings/drawing1.xml'
# rc = []
# if xml_name in fz.namelist():
#     xml_string = fz.read(xml_name).decode()
#     images_row_numbers = re.findall(r'<xdr:row>(\d+)</xdr:row>', xml_string)
#     images_col_numbers = re.findall(r'<xdr:col>(\d+)</xdr:col>', xml_string)
#     embeds = re.findall(r'<a:blip r:embed="rId(\d+)"/', xml_string)
#     # print(embeds)
#     if images_row_numbers and images_col_numbers:

#         res = list(map(int, images_row_numbers))
#         clos = list(map(int, images_col_numbers))
#         embed = list(map(int, embeds))
# fz.close()
# res = [r + 1 for r in res]
# rc = zip(res,clos,embed)
# print(list(rc))




